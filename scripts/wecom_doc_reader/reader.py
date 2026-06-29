#!/usr/bin/env python3
"""WeComDocReader — 企微文档浏览器读取主类"""

from __future__ import annotations

import asyncio
import json
import os
import re
import sys
import time
import zlib
import base64
import fcntl
from pathlib import Path
from urllib.parse import urlparse, parse_qs
from datetime import datetime, timezone, timedelta

from .constants import (
    __version__,
    DEFAULT_STATE_DIR,
    LEGACY_SHARED_STATE,
    _TZ_CN,
    _URL_PREFIX_TO_TYPE,
    _FIELD_TYPE_TEXT,
    _FIELD_TYPE_USER,
    _FIELD_TYPE_EDITOR,
    _FIELD_TYPE_CREATED_AT,
    _FIELD_TYPE_UPDATED_AT,
    _FIELD_TYPE_SELECT,
    _FIELD_TYPE_FORMULA,
    _DOM_SELECTORS,
)
from .utils import _auto_report
from .parsers import (
    parse_doc_url,
    _cell_value,
    _parse_column_defs,
    _block_fonts,
    _is_login_page,
)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 主类
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class WeComDocReader:
    """
    企微文档浏览器读取器（用户隔离版）

    每个 user_id 拥有独立的 storage_state 文件。
    通过文件锁防止同一用户并发登录。
    """

    def __init__(self, state_dir: str = DEFAULT_STATE_DIR):
        self.state_dir = Path(state_dir)
        self.state_dir.mkdir(parents=True, exist_ok=True)

    # ── 内部路径 ──

    def _safe_id(self, user_id: str) -> str:
        return re.sub(r"[^\w\-.]", "_", user_id)

    def _state_file(self, user_id: str) -> Path:
        return self.state_dir / f"{self._safe_id(user_id)}.json"

    def _lock_file(self, user_id: str) -> Path:
        return self.state_dir / f"{self._safe_id(user_id)}.lock"

    def _qr_image(self, user_id: str) -> Path:
        return self.state_dir / f"{self._safe_id(user_id)}_qr.png"

    def _has_state(self, user_id: str) -> bool:
        return self._state_file(user_id).exists()

    def _acquire_lock(self, user_id: str):
        """获取用户级文件锁（防止并发登录）"""
        lock_path = self._lock_file(user_id)
        fd = open(lock_path, "w")
        try:
            fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
            fd.write(str(os.getpid()))
            fd.flush()
            return fd
        except BlockingIOError:
            fd.close()
            return None

    def _release_lock(self, fd):
        if fd:
            fcntl.flock(fd, fcntl.LOCK_UN)
            fd.close()

    # ── 1. 扫码登录 ──────────────────────────────────────

    async def login(self, user_id: str, timeout: int = 300) -> dict:
        """
        为指定用户启动扫码登录。

        返回:
            {"ok": True, "cookies": N, "state_file": "...", "qr_image": "..."}
            {"ok": False, "error": "...", "qr_image": "..."}
        """
        from playwright.async_api import async_playwright

        # 并发保护
        lock_fd = self._acquire_lock(user_id)
        if not lock_fd:
            return {"ok": False, "error": f"用户 {user_id} 正在登录中（并发锁冲突）"}

        state_file = str(self._state_file(user_id))
        qr_image = str(self._qr_image(user_id))

        try:
            async with async_playwright() as p:
                browser = await p.chromium.launch(
                    headless=True, args=["--no-sandbox", "--disable-setuid-sandbox"]
                )
                # 不传旧 state，强制显示登录二维码
                context = await browser.new_context(
                    viewport={"width": 800, "height": 600}
                )
                page = await context.new_page()
                await _block_fonts(page)

                # 捕获二维码图片
                qr_captured = False

                async def capture_qr(response):
                    nonlocal qr_captured
                    if "qrcode" in response.url.lower() and response.status == 200:
                        ct = response.headers.get("content-type", "")
                        if "image" in ct and not qr_captured:
                            try:
                                data = await response.body()
                                with open(qr_image, "wb") as f:
                                    f.write(data)
                                qr_captured = True
                            except:
                                pass

                page.on("response", capture_qr)

                await page.goto(
                    "https://doc.weixin.qq.com",
                    wait_until="domcontentloaded", timeout=15000,
                )

                # 等待 QR 码出现（最多 20 秒）
                for _ in range(20):
                    await asyncio.sleep(1)
                    if qr_captured:
                        break

                if not qr_captured:
                    await browser.close()
                    return {
                        "ok": False,
                        "error": "二维码未能加载，请检查网络或 Playwright 安装",
                        "qr_image": qr_image,
                    }

                # 等待扫码（最多 timeout 秒）
                for _ in range(timeout // 2):
                    await asyncio.sleep(2)
                    if not _is_login_page(page.url):
                        # 扫码成功，等待页面加载
                        await asyncio.sleep(5)
                        state = await context.storage_state()
                        with open(state_file, "w") as f:
                            json.dump(state, f, ensure_ascii=False, indent=2)
                        n = len(state.get("cookies", []))
                        await browser.close()
                        return {
                            "ok": True,
                            "cookies": n,
                            "state_file": state_file,
                            "qr_image": qr_image,
                        }

                await browser.close()
                return {
                    "ok": False,
                    "error": f"等待超时（{timeout}s），二维码可能已过期",
                    "qr_image": qr_image,
                }
        finally:
            self._release_lock(lock_fd)

    # ── 2. Cookies 检查 ──────────────────────────────────

    async def check(self, user_id: str) -> dict:
        """
        检查指定用户的 cookies 是否有效。

        返回:
            {"valid": bool, "user_id": str, "remaining_days": float, "reason": str}
        """
        from playwright.async_api import async_playwright

        state_file = str(self._state_file(user_id))
        if not self._has_state(user_id):
            return {"valid": False, "user_id": user_id, "reason": "未登录（无 storage_state）"}

        # 静态检查：过期时间
        remaining_days = None
        try:
            with open(state_file) as f:
                state = json.load(f)
            for c in state.get("cookies", []):
                if c["name"] == "wedoc_sid":
                    remaining_days = (c.get("expires", 0) - time.time()) / 86400
        except:
            pass

        if remaining_days is not None and remaining_days < 0:
            return {
                "valid": False,
                "user_id": user_id,
                "remaining_days": round(remaining_days, 1),
                "reason": "cookies 已过期（静态检查）",
            }

        # 动态检查：打开页面验证
        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=True, args=["--no-sandbox", "--disable-setuid-sandbox"]
            )
            ctx = await browser.new_context(
                storage_state=state_file,
                viewport={"width": 800, "height": 600},
            )
            page = await ctx.new_page()
            await _block_fonts(page)

            valid = False
            url = ""
            try:
                await page.goto(
                    "https://doc.weixin.qq.com/home/recent",
                    wait_until="domcontentloaded", timeout=15000,
                )
                await asyncio.sleep(3)
                url = page.url
                body = await page.evaluate(
                    "() => document.body?.innerText?.substring(0, 200) || ''"
                )
                is_login = _is_login_page(url)
                has_content = any(k in body for k in ["我的文档", "首页", "最近查看"])
                valid = not is_login and has_content
            except Exception as e:
                valid = False
                url = str(e)[:100]
            await browser.close()

        result = {"valid": valid, "user_id": user_id}
        if remaining_days is not None:
            result["remaining_days"] = round(remaining_days, 1)
        result["reason"] = "OK" if valid else f"页面跳转到: {url}"
        return result

    # ── 3. 读取文档（自动路由） ──────────────────────────

    async def read(self, user_id: str, url: str, sheet_id: str = None) -> dict:
        """
        读取企微文档。自动根据 URL 路由:
          - s3_ 智能表格 → _read_smartsheet (dop-api 全量结构化)
          - e3_ 电子表格 → _read_spreadsheet (原生 JS API + 剪贴板 HTML 兜底)
          - m4_ 思维导图 → _read_mind (dop-api/get/mind JSON 节点树)
          - 其他类型     → _read_dom (DOM 文本提取)

        返回:
            成功: {"success": True, "doc_type": "...", ...}
            失败: {"success": False, "error": "..."}
        """
        if not self._has_state(user_id):
            return {
                "success": False,
                "error": f"用户 {user_id} 未登录，请先调用 login()",
                "user_id": user_id,
            }

        info = parse_doc_url(url)

        if info["doc_type"] == "s3":
            return await self._read_smartsheet(
                user_id, url, sheet_id or info["tab"]
            )
        elif info["doc_type"] == "e3":
            return await self._read_spreadsheet(
                user_id, url, sheet_id or info["tab"]
            )
        elif info["doc_type"] == "m4":
            return await self._read_mind(user_id, url, info)
        else:
            return await self._read_dom(user_id, url, info)

    # ── 3a. DOM 文本提取（通用） ─────────────────────────

    async def _read_dom(self, user_id: str, url: str, info: dict) -> dict:
        """用 Playwright 打开页面，从 DOM 提取文本"""
        from playwright.async_api import async_playwright

        state_file = str(self._state_file(user_id))

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=True, args=["--no-sandbox", "--disable-setuid-sandbox"]
            )
            ctx = await browser.new_context(
                storage_state=state_file,
                viewport={"width": 1280, "height": 800},
            )
            page = await ctx.new_page()
            await _block_fonts(page)

            try:
                await page.goto(url, wait_until="networkidle", timeout=30000)
            except Exception as e:
                # 超时但页面可能已加载
                pass

            if _is_login_page(page.url):
                await browser.close()
                return {
                    "success": False,
                    "error": "cookies 已过期，需重新登录",
                    "user_id": user_id,
                }

            # 检查是否有权限
            body_preview = await page.evaluate(
                "() => document.body?.innerText?.substring(0, 300) || ''"
            )
            if "暂无权限" in body_preview or "申请权限" in body_preview:
                await browser.close()
                return {
                    "success": False,
                    "error": "当前用户无此文档的访问权限",
                    "user_id": user_id,
                    "hint": "需要文档所有者授权，或用有权限的账号重新扫码",
                }

            await asyncio.sleep(3)

            content = await page.evaluate(_DOM_SELECTORS)
            title = await page.title()

            await browser.close()

        return {
            "success": content.get("success", False),
            "doc_type": info["doc_type"],
            "path_type": info["path_type"],
            "title": title,
            "url": url,
            "selector": content.get("selector"),
            "text_length": len(content.get("text", "")),
            "text": content.get("text", ""),
        }

    # ── 3b. 电子表格（e3_）增强读取 ──────────────────────

    async def _read_spreadsheet(
        self, user_id: str, url: str, sheet_id: str = None
    ) -> dict:
        """
        读取 e3_ 电子表格。优化策略（v3.0.0 — 基于实测）:
          1. JS Runtime 元数据提取 (sheet 列表/名称/合并单元格/图片URL)
          2. 剪贴板 HTML 多子表遍历 (含 colspan/rowspan 合并单元格保留)
          3. xlsx 导出 + openpyxl 解析
          4. 剪贴板 TSV 纯文本兜底
          5. DOM 文本提取（最终兜底）

        实测结论:
          - dop-api 对 e3_ 返回 protobuf 二进制格式，无法直接 JSON.parse
          - JS 运行时 SpreadsheetApp.workbook 提供元数据但 cell 数据懒加载
          - 剪贴板 HTML 是当前唯一稳定获取完整 cell 数据的方式
        """
        from playwright.async_api import async_playwright

        state_file = str(self._state_file(user_id))
        info = parse_doc_url(url)

        if not info["doc_id"]:
            return {"success": False, "error": f"无法解析 URL: {url}"}

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=True, args=["--no-sandbox", "--disable-setuid-sandbox"]
            )
            ctx = await browser.new_context(
                storage_state=state_file,
                viewport={"width": 1920, "height": 1080},
                permissions=["clipboard-read", "clipboard-write"],
            )
            page = await ctx.new_page()
            await _block_fonts(page)

            try:
                await page.goto(url, wait_until="networkidle", timeout=60000)
            except:
                pass
            await page.wait_for_timeout(8000)  # 等待 JS 运行时初始化

            if _is_login_page(page.url):
                await browser.close()
                return {
                    "success": False,
                    "error": "cookies 已过期，需重新登录",
                    "user_id": user_id,
                }

            body_preview = await page.evaluate(
                "() => document.body?.innerText?.substring(0, 300) || ''"
            )
            if "暂无权限" in body_preview or "申请权限" in body_preview:
                await browser.close()
                return {
                    "success": False,
                    "error": "当前用户无此文档的访问权限",
                    "user_id": user_id,
                }

            title = await page.title()

            # ── Phase 1: JS Runtime 元数据 ──
            js_meta = await self._get_js_runtime_sheets(page)
            sheet_count = js_meta.get("sheetCount", 0) if js_meta else 0
            if js_meta:
                print(f"  JS Runtime: {sheet_count} sheets")

            # ── Phase 2: 原生 JS API 读取（getCellDataAtPosition）──
            native_result = await self._read_all_sheets_via_native_api(page, js_meta)
            if native_result and native_result.get("success"):
                await browser.close()
                native_result["title"] = title
                native_result["method"] = "native-js-api"
                return native_result

            # ── Phase 3: xlsx 导出（需要编辑权限）──
            xlsx_result = await self._try_xlsx_export_for_spreadsheet(page)
            if xlsx_result and xlsx_result.get("success"):
                await browser.close()
                xlsx_result["title"] = title
                return xlsx_result

            # ── Phase 4: 剪贴板 HTML 兜底 ──
            html_result = await self._try_clipboard_html_all_sheets(page)
            if html_result and html_result.get("success"):
                await browser.close()
                html_result["title"] = title
                html_result["method"] = "clipboard-html"
                return html_result

            # ── Phase 5: DOM 兜底 ──
            content = await page.evaluate(_DOM_SELECTORS)
            await browser.close()

            return {
                "success": content.get("success", False),
                "doc_type": "e3",
                "path_type": "sheet",
                "title": title,
                "url": url,
                "method": "dom-fallback",
                "selector": content.get("selector"),
                "text_length": len(content.get("text", "")),
                "text": content.get("text", ""),
                "warning": "仅获取到可见区域文本，单元格数据可能不完整",
            }

    async def _get_js_runtime_sheets(self, page) -> dict:
        """从 SpreadsheetApp.workbook JS 运行时提取所有 sheet 的元数据。
        
        返回:
            {sheet_id: {name, mergeList, images, usedRange}, ...}
            失败返回空 dict
        """
        try:
            result = await page.evaluate('''() => {
                try {
                    const wb = window.SpreadsheetApp?.workbook;
                    if (!wb) return {error: 'no SpreadsheetApp.workbook'};
                    const wm = wb.worksheetManager;
                    if (!wm?.sheetList) return {error: 'no sheetList'};
                    
                    const sheets = {};
                    for (const sheet of wm.sheetList) {
                        const ur = sheet.cellDataGrid?.usedRange;
                        const sid = ur?.sheetId;
                        if (!sid) continue;
                        
                        let name = sid;
                        try { name = wm.getSheetNameBySheetId(sid); } catch(e) {}
                        
                        // 合并单元格
                        const merges = sheet.mergeManager?.mergeList;
                        const mergeList = Array.isArray(merges) ? merges.map(m => {
                            if (m && typeof m === 'object') {
                                return {
                                    startRow: m.startRowIndex ?? m.startRow ?? m.row,
                                    startCol: m.startColIndex ?? m.startCol ?? m.col,
                                    endRow: m.endRowIndex ?? m.endRow,
                                    endCol: m.endColIndex ?? m.endCol,
                                };
                            }
                            return m;
                        }) : [];
                        
                        sheets[sid] = {
                            name: name,
                            sheetId: sid,
                            mergeList: mergeList,
                            mergeCount: mergeList.length,
                            usedRange: ur ? {
                                startRow: ur.startRowIndex,
                                endRow: ur.endRowIndex,
                                startCol: ur.startColIndex,
                                endCol: ur.endColIndex,
                            } : null,
                        };
                    }
                    
                    // 图片数据
                    const dm = wb.drawingManager;
                    let images = {};
                    if (dm?.drawingMap) {
                        for (const [imgId, drawing] of Object.entries(dm.drawingMap)) {
                            images[imgId] = {id: imgId};
                            // 递归搜索 URL
                            function findUrls(obj, depth) {
                                if (depth > 5) return;
                                if (typeof obj === 'string' && (obj.startsWith('https://') || obj.startsWith('http://'))) {
                                    images[imgId].url = obj;
                                }
                                if (typeof obj === 'object') {
                                    for (const k of Object.keys(obj)) findUrls(obj[k], depth + 1);
                                }
                            }
                            findUrls(drawing, 0);
                        }
                    }
                    
                    return {sheets: sheets, images: images, sheetCount: Object.keys(sheets).length};
                } catch(e) {
                    return {error: e.message};
                }
            }''')
            
            if result.get("error"):
                return {}
            return result
        except Exception:
            return {}

    async def _read_all_sheets_via_native_api(self, page, js_meta: dict) -> dict:
        """
        使用企微表格引擎的原生 JS API (getCellDataAtPosition) 读取所有子表。
        
        这是最直接、最稳定的方式：
        - 直接调用表格引擎的内部 API，无需剪贴板/导出等间接方式
        - 支持精确读取合并单元格范围 (getMergeReference)
        - 支持图片列原始 URL (getExtendedValue)
        - 性能极高（800 cells < 1ms）
        
        限制：非活跃 tab 需要切换后等待数据加载（懒加载）。
        """
        if not js_meta or not js_meta.get("sheets"):
            return {"success": False, "error": "JS Runtime 元数据不可用"}
        
        sheets_info = js_meta["sheets"]  # {sheet_id: {name, sheetId, ...}}
        sheet_list = list(sheets_info.values())
        
        if not sheet_list:
            return {"success": False, "error": "未检测到子表"}
        
        all_sheets = {}
        all_records = []
        failed_sheets = []
        
        for idx, sheet_info in enumerate(sheet_list):
            sheet_name = sheet_info.get("name", "unknown")
            sheet_id = sheet_info.get("sheetId", "unknown")
            print(f"  [{idx+1}/{len(sheet_list)}] {sheet_name} ({sheet_id})")
            
            # 点击 tab 切换到目标子表
            tab_els = await page.query_selector_all('.tab-bar-item-title')
            clicked = False
            for el in tab_els:
                t = await el.text_content()
                if t and t.strip() == sheet_name:
                    await el.scroll_into_view_if_needed()
                    await page.wait_for_timeout(300)
                    await el.click()
                    clicked = True
                    break
            
            if not clicked:
                print(f"    ✗ 无法点击 tab")
                failed_sheets.append(sheet_name)
                continue
            
            # 等待数据加载（懒加载，切换 tab 后需要时间）
            await page.wait_for_timeout(4000)
            
            # 用 JS 批量读取所有 cell
            js_result = await page.evaluate('''(sheetId) => {
                try {
                    const wb = window.SpreadsheetApp?.workbook;
                    if (!wb) return {error: "no workbook"};
                    const wm = wb.worksheetManager;
                    
                    // 找目标 sheet
                    let target = null;
                    for (const s of wm.sheetList) {
                        if (s.cellDataGrid?.usedRange?.sheetId === sheetId) {
                            target = s;
                            break;
                        }
                    }
                    if (!target) return {error: "sheet not found: " + sheetId};
                    
                    const ur = target.cellDataGrid.usedRange;
                    const startRow = ur.startRowIndex;
                    const endRow = ur.endRowIndex;
                    const startCol = ur.startColIndex;
                    const endCol = ur.endColIndex;
                    
                    // 边界检查
                    if (startRow > endRow || startCol > endCol) {
                        return {error: "empty sheet", usedRange: ur};
                    }
                    
                    // 限制最大读取范围（防止超大表格卡死）
                    const maxRows = Math.min(endRow - startRow + 1, 500);
                    const maxCols = Math.min(endCol - startCol + 1, 50);
                    
                    const rows = [];
                    const merges = [];
                    const imageUrls = {};  // {row_col: url}
                    
                    for (let r = 0; r < maxRows; r++) {
                        const row = [];
                        for (let c = 0; c < maxCols; c++) {
                            const cell = target.getCellDataAtPosition(startRow + r, startCol + c);
                            if (!cell || cell.isEmpty()) {
                                row.push(null);
                                continue;
                            }
                            
                            const cellType = cell.getType();
                            let value = cell.getValue();
                            
                            // 公式/复杂类型：用 formattedValue 拿纯文本
                            if (value && typeof value === "object" && !Array.isArray(value)) {
                                try {
                                    const fv = cell.getFormattedValue();
                                    if (fv && fv.value) value = fv.value;
                                } catch(e) {}
                            }
                            
                            // 图片列：getExtendedValue() 返回 JSON 字符串 ["url", w, h]
                            let extType = null;
                            let extValue = null;
                            try {
                                extType = cell.getExtendedType();
                                if (extType === 5) {  // 5 = image/attachment
                                    extValue = cell.getExtendedValue();
                                    if (extValue && typeof extValue === "string") {
                                        try {
                                            const parsed = JSON.parse(extValue);
                                            if (Array.isArray(parsed) && parsed[0]) {
                                                imageUrls[r + "_" + c] = parsed[0];
                                            }
                                        } catch(e) {}
                                    }
                                }
                            } catch(e) {}
                            
                            // 合并单元格：getMergeReference() 返回精确范围
                            let mergeRef = null;
                            try {
                                const mr = cell.getMergeReference();
                                if (mr) {
                                    mergeRef = {
                                        startRow: mr.startRowIndex,
                                        endRow: mr.endRowIndex,
                                        startCol: mr.startColIndex,
                                        endCol: mr.endColIndex,
                                    };
                                    // 只记录合并区域的左上角
                                    if (mergeRef.startRow === startRow + r && mergeRef.startCol === startCol + c) {
                                        merges.push(mergeRef);
                                    }
                                }
                            } catch(e) {}
                            
                            row.push({
                                value: value,
                                type: cellType,
                                mergeRef: mergeRef,
                            });
                        }
                        rows.push(row);
                    }
                    
                    return {
                        success: true,
                        usedRange: {startRow, endRow, startCol, endCol},
                        readRange: {maxRows, maxCols},
                        rows: rows,
                        merges: merges,
                        imageUrls: imageUrls,
                        rowCount: rows.length,
                        colCount: maxCols,
                    };
                } catch(e) {
                    return {error: e.message};
                }
            }''', sheet_id)
            
            if not js_result or js_result.get("error"):
                err = js_result.get("error", "unknown") if js_result else "null result"
                print(f"    ✗ 读取失败: {err}")
                failed_sheets.append(sheet_name)
                continue
            
            # Python 端处理：转换为 records
            raw_rows = js_result.get("rows", [])
            image_urls = js_result.get("imageUrls", {})
            merge_list = js_result.get("merges", [])
            
            if not raw_rows:
                print(f"    ⚠ 空表")
                all_sheets[sheet_name] = {
                    "rows": [], "row_count": 0, "headers": [],
                    "sheetId": sheet_id, "mergeList": merge_list,
                    "mergeCount": len(merge_list), "method": "native-js-api",
                }
                continue
            
            # 第 0 行作为表头
            headers = []
            for c, cell_data in enumerate(raw_rows[0]):
                if cell_data is None:
                    headers.append(f"col_{c}")
                elif isinstance(cell_data, dict):
                    v = cell_data.get("value")
                    if v is None:
                        headers.append(f"col_{c}")
                    elif isinstance(v, (int, float)) and 30000 < v < 100000:
                        # 表头中的日期也转换
                        try:
                            from datetime import datetime, timedelta
                            base = datetime(1899, 12, 30)
                            actual_date = base + timedelta(days=int(v))
                            headers.append(actual_date.strftime("%-m月%-d日"))
                        except:
                            headers.append(str(v))
                    else:
                        headers.append(str(v))
                else:
                    headers.append(str(cell_data))
            
            # 数据行
            records = []
            for r_idx, raw_row in enumerate(raw_rows[1:], start=1):
                row_dict = {"_row": r_idx}
                for c_idx, cell_data in enumerate(raw_row):
                    key = headers[c_idx] if c_idx < len(headers) else f"col_{c_idx}"
                    
                    if cell_data is None:
                        row_dict[key] = ""
                        continue
                    
                    if isinstance(cell_data, dict):
                        value = cell_data.get("value")
                        cell_type = cell_data.get("type")
                        
                        # 日期转换：Excel serial number → ISO 日期
                        if isinstance(value, (int, float)) and 30000 < value < 100000:
                            # Excel 日期基准: 1900-01-01 = 1, 但有 1900-02-29 bug
                            try:
                                from datetime import datetime, timedelta
                                base = datetime(1899, 12, 30)  # Excel epoch
                                actual_date = base + timedelta(days=int(value))
                                value = actual_date.strftime("%Y-%m-%d")
                            except:
                                pass
                        
                        row_dict[key] = value if value is not None else ""
                    else:
                        row_dict[key] = cell_data
                    
                    # 图片 URL
                    img_key = f"{r_idx}_{c_idx}"
                    if img_key in image_urls:
                        row_dict[f"{key}_图片URL"] = image_urls[img_key]
                
                row_dict["_sheet_name"] = sheet_name
                records.append(row_dict)
            
            # ── 合并单元格填充 ──
            # mergeList 行号是 sheet 级别（0=表头），records 是数据级别（0=第一条数据）
            # 所以 record_index = merge_row - 1
            for m in merge_list:
                sr = m.get("startRow", 0)
                er = m.get("endRow", 0)
                sc = m.get("startCol", 0)
                ec = m.get("endCol", 0)
                
                # 获取合并起始格的 header
                start_header = headers[sc] if sc < len(headers) else None
                if not start_header:
                    continue
                
                # 获取起始格的值（sheet row sr → record index sr-1）
                start_record_idx = sr - 1  # -1 because records[0] = sheet row 1
                if start_record_idx < 0 or start_record_idx >= len(records):
                    continue
                
                fill_value = records[start_record_idx].get(start_header, "")
                
                # 填充合并范围内的所有格
                for r in range(sr, er + 1):
                    rec_idx = r - 1
                    if rec_idx < 0 or rec_idx >= len(records):
                        continue
                    for c in range(sc, ec + 1):
                        h = headers[c] if c < len(headers) else None
                        if not h:
                            continue
                        current = records[rec_idx].get(h, "")
                        if current == "" or current is None:
                            records[rec_idx][h] = fill_value
            
            all_sheets[sheet_name] = {
                "rows": records,
                "row_count": len(records),
                "headers": headers,
                "sheetId": sheet_id,
                "mergeList": merge_list,
                "mergeCount": len(merge_list),
                "method": "native-js-api",
                "usedRange": js_result.get("usedRange"),
                "readRange": js_result.get("readRange"),
            }
            all_records.extend(records)
            print(f"    ✓ {len(records)} rows, {len(merge_list)} merges, "
                  f"{len(image_urls)} images")
        
        if not all_sheets:
            return {"success": False, "error": f"所有 {len(sheet_list)} 个子表读取失败"}
        
        return {
            "success": True,
            "doc_type": "e3",
            "path_type": "sheet",
            "url": page.url,
            "sheets": all_sheets,
            "sheet_names": list(all_sheets.keys()),
            "sheet_count": len(all_sheets),
            "total": len(all_records),
            "records": all_records,
            "failed_sheets": failed_sheets,
        }

    async def _do_ctrl_a_c(self, page) -> bool:
        """执行 Ctrl+A 全选 + Ctrl+C 复制的公共操作。返回是否成功点击了 canvas。"""
        # 隐藏遮罩层
        await page.evaluate('''() => {
            document.querySelectorAll(".operate-board, .sheet-mask, .toolbar-mask")
                .forEach(el => el.style.display = "none");
        }''')
        # 用 mouse.click 点击 canvas 区域中心
        viewport = page.viewport_size or {"width": 1920, "height": 1080}
        await page.mouse.click(viewport["width"] // 2, viewport["height"] // 2)
        await page.wait_for_timeout(500)
        # Ctrl+A + Ctrl+C
        await page.keyboard.press("Control+a")
        await page.wait_for_timeout(500)
        await page.keyboard.press("Control+c")
        await page.wait_for_timeout(1500)
        return True

    async def _try_clipboard_html_all_sheets(self, page) -> dict:
        """多子表遍历 + 剪贴板 HTML 提取。逐个切换 tab，对每个子表做 Ctrl+A/C → 读 HTML clipboard。"""
        try:
            # 获取所有子表名
            tab_els = await page.query_selector_all('.tab-bar-item-title')
            tab_names = []
            for el in tab_els:
                t = await el.text_content()
                if t and t.strip():
                    tab_names.append(t.strip())

            if not tab_names:
                return {"success": False, "error": "未检测到子表"}

            all_sheets = {}
            all_records = []
            total_merged = 0
            failed_sheets = []

            for idx, tab_name in enumerate(tab_names):
                # 重新获取 tab 元素（切换后 DOM 可能变化）
                tab_els = await page.query_selector_all('.tab-bar-item-title')
                clicked = False
                for el in tab_els:
                    t = await el.text_content()
                    if t and t.strip() == tab_name:
                        await el.scroll_into_view_if_needed()
                        await page.wait_for_timeout(300)
                        await el.click()
                        clicked = True
                        break
                if not clicked:
                    failed_sheets.append(tab_name)
                    continue

                # 等待渲染（5s，确保 canvas 完全重绘）
                await page.wait_for_timeout(5000)

                # 重新隐藏遮罩（每次切 tab 后遮罩会重新出现）
                await self._do_ctrl_a_c(page)

                # 读剪贴板 HTML
                html_content = await page.evaluate("""async () => {
                    try {
                        const items = await navigator.clipboard.read();
                        for (const item of items) {
                            if (item.types.includes('text/html')) {
                                const blob = await item.getType('text/html');
                                return await blob.text();
                            }
                        }
                        return null;
                    } catch(e) { return null; }
                }""")

                if not html_content or len(html_content) < 20:
                    # HTML 不可用，尝试纯文本
                    tsv_text = await page.evaluate("""async () => {
                        try { return await navigator.clipboard.readText(); }
                        catch(e) { return null; }
                    }""")
                    if tsv_text and len(tsv_text) > 10:
                        lines = tsv_text.strip().split("\n")
                        headers = lines[0].split("\t") if lines else []
                        rows = []
                        for line in lines[1:]:
                            cells = line.split("\t")
                            row = {}
                            for i, cell in enumerate(cells):
                                key = headers[i] if i < len(headers) else f"col_{i}"
                                row[key] = cell.strip()
                            row["_sheet_name"] = tab_name
                            rows.append(row)
                        all_sheets[tab_name] = {
                            "rows": rows, "row_count": len(rows),
                            "headers": headers, "method": "tsv-fallback",
                            "merged_ranges": [], "merged_cell_count": 0,
                        }
                        all_records.extend(rows)
                    else:
                        failed_sheets.append(tab_name)
                    continue

                # 解析 HTML table
                parsed = self._parse_html_table(html_content)
                if not parsed or not parsed.get("records"):
                    # HTML 有内容但无 table → 可能是 canvas 渲染导致 Ctrl+A/C 只复制了图片
                    # 改用页面截图 + VL OCR 识别
                    has_img = bool(html_content and '<img' in html_content)
                    if has_img or (html_content and '<table' not in html_content):
                        print(f"  [{idx+1}/{len(tab_names)}] {tab_name}: 剪贴板无表格数据，用截图+VL OCR...")
                        # 隐藏工具栏再截图
                        await page.evaluate('''() => {
                            document.querySelectorAll(".toolbar, .top-bar, .tab-bar, .sidebar, .bottom-bar, .footer")
                                .forEach(el => el.style.display = "none");
                        }''')
                        await page.wait_for_timeout(300)
                        import base64 as _b64
                        ss_bytes = await page.screenshot(full_page=False)
                        img_b64 = _b64.b64encode(ss_bytes).decode()
                        vl_result = await self._try_vl_ocr_for_image(img_b64, "image/png")
                        if vl_result.get("success"):
                            for rec in vl_result["records"]:
                                rec["_sheet_name"] = tab_name
                            all_sheets[tab_name] = {
                                "rows": vl_result["records"],
                                "row_count": len(vl_result["records"]),
                                "headers": vl_result.get("headers", []),
                                "merged_ranges": [],
                                "merged_cell_count": 0,
                                "method": "vl-ocr",
                            }
                            all_records.extend(vl_result["records"])
                            print(f"  [{idx+1}/{len(tab_names)}] {tab_name}: VL OCR 成功，{len(vl_result['records'])} 行")
                        else:
                            print(f"  [{idx+1}/{len(tab_names)}] {tab_name}: VL OCR 失败 - {vl_result.get('error')}")
                            failed_sheets.append(tab_name)
                    else:
                        failed_sheets.append(tab_name)
                    continue

                # 给每条记录加上子表标识
                for rec in parsed["records"]:
                    rec["_sheet_name"] = tab_name

                all_sheets[tab_name] = {
                    "rows": parsed["records"],
                    "row_count": len(parsed["records"]),
                    "headers": parsed.get("headers", []),
                    "merged_ranges": parsed.get("merged_ranges", []),
                    "merged_cell_count": parsed.get("merged_cell_count", 0),
                    "method": "html",
                }
                all_records.extend(parsed["records"])
                total_merged += parsed.get("merged_cell_count", 0)

            if not all_sheets:
                return {"success": False, "error": f"所有 {len(tab_names)} 个子表提取失败"}

            return {
                "success": True,
                "doc_type": "e3",
                "path_type": "sheet",
                "url": page.url,
                "sheets": all_sheets,
                "sheet_names": list(all_sheets.keys()),
                "sheet_count": len(all_sheets),
                "total": len(all_records),
                "records": all_records,
                "merged_cell_count": total_merged,
                "failed_sheets": failed_sheets,
            }
        except Exception as e:
            return {"success": False, "error": f"多子表 HTML 提取失败: {e}"}

    def _parse_html_table(self, html_content: str) -> dict:
        """解析剪贴板 HTML 中的 table，返回 records + headers + merged_cell_count。公共方法，供单表/多表共用。"""
        from html.parser import HTMLParser

        class TableParser(HTMLParser):
            def __init__(self):
                super().__init__()
                self.rows = []
                self.current_row = []
                self.current_cell = {"text": "", "colspan": 1, "rowspan": 1}
                self.in_td = False
                self.in_table = False

            def handle_starttag(self, tag, attrs):
                attrs_dict = dict(attrs)
                if tag == "table":
                    self.in_table = True
                elif tag in ("td", "th") and self.in_table:
                    self.in_td = True
                    self.current_cell = {
                        "text": "",
                        "colspan": int(attrs_dict.get("colspan", 1)),
                        "rowspan": int(attrs_dict.get("rowspan", 1)),
                    }
                elif tag == "tr" and self.in_table:
                    self.current_row = []
                elif tag == "br" and self.in_td:
                    self.current_cell["text"] += "\n"

            def handle_endtag(self, tag):
                if tag in ("td", "th") and self.in_td:
                    self.in_td = False
                    self.current_row.append(self.current_cell)
                elif tag == "tr" and self.in_table:
                    if self.current_row:
                        self.rows.append(self.current_row)
                    self.current_row = []
                elif tag == "table":
                    self.in_table = False

            def handle_data(self, data):
                if self.in_td:
                    self.current_cell["text"] += data

        parser = TableParser()
        parser.feed(html_content)

        if not parser.rows:
            return None

        # 展开合并单元格为二维 grid
        grid = {}
        max_col = 0
        for r, row_cells in enumerate(parser.rows):
            col_offset = 0
            for cell in row_cells:
                while (r, col_offset) in grid:
                    col_offset += 1
                text = cell["text"].strip()
                cs, rs = cell["colspan"], cell["rowspan"]
                for dr in range(rs):
                    for dc in range(cs):
                        grid[(r + dr, col_offset + dc)] = text
                col_offset += cs
                max_col = max(max_col, col_offset)

        if not grid:
            return None

        # 第 0 行作为表头
        headers = [grid.get((0, c), f"col_{c}") for c in range(max_col)]

        # 数据行
        max_row = max(r for r, _ in grid.keys()) if grid else 0
        records = []
        for r in range(1, max_row + 1):
            row = {}
            for c in range(max_col):
                key = headers[c] if c < len(headers) else f"col_{c}"
                row[key] = grid.get((r, c), "")
            records.append(row)

        # 合并单元格统计
        merged_count = sum(
            1 for row in parser.rows for cell in row
            if cell["colspan"] > 1 or cell["rowspan"] > 1
        )

        return {
            "headers": headers,
            "records": records,
            "merged_cell_count": merged_count,
        }

    async def _try_vl_ocr_for_image(self, img_base64: str, mime: str = "image/png") -> dict:
        """用 qwen-vl-max 识别图片中的表格内容，返回结构化记录。
        
        适用场景：子表内容全是图片（如宣传工作日志），剪贴板只拿到 base64 图片。
        """
        try:
            import requests
            # 读 dashscope key
            key_path = "/tmp/dashscope_key.txt"
            if not os.path.exists(key_path):
                return {"success": False, "error": "dashscope key 不存在"}
            with open(key_path) as f:
                api_key = f.read().strip()

            prompt = """这是一张电子表格的图片。请完整识别图中所有文字和数字，按表格结构输出。
要求：
1. 逐字读取所有可见文字，不要遗漏
2. 如果有表头，第一行输出表头
3. 每行数据用 | 分隔
4. 保持原始列顺序
5. 如果有合并单元格，重复对应的值
6. 只输出表格内容，不要加任何解释文字
7. 用 markdown 表格格式输出"""

            url = "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions"
            headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
            payload = {
                "model": "qwen-vl-max",
                "messages": [{"role": "user", "content": [
                    {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{img_base64}"}},
                    {"type": "text", "text": prompt}
                ]}],
                "max_tokens": 4000
            }

            resp = requests.post(url, headers=headers, json=payload, timeout=60)
            result = resp.json()
            if "choices" not in result:
                return {"success": False, "error": f"VL API 错误: {result.get('message', 'unknown')}"}

            vl_text = result["choices"][0]["message"]["content"]

            # 解析 markdown 表格
            lines = [l.strip() for l in vl_text.strip().split("\n") if l.strip()]
            # 去掉 markdown 分隔行 (|---|---|)
            table_lines = [l for l in lines if not (l.startswith("|") and set(l.replace("|", "").replace("-", "").replace(":", "").strip()) == set())]

            if len(table_lines) < 2:
                return {"success": False, "error": f"VL 输出不是表格: {vl_text[:200]}"}

            # 解析
            headers_row = [c.strip() for c in table_lines[0].strip("|").split("|")]
            records = []
            for line in table_lines[1:]:
                cells = [c.strip() for c in line.strip("|").split("|")]
                row = {}
                for i, cell in enumerate(cells):
                    key = headers_row[i] if i < len(headers_row) else f"col_{i}"
                    row[key] = cell
                records.append(row)

            return {
                "success": True,
                "headers": headers_row,
                "records": records,
                "vl_raw": vl_text,
            }
        except Exception as e:
            return {"success": False, "error": f"VL OCR 失败: {e}"}

    async def _try_clipboard_html_for_spreadsheet(self, page) -> dict:
        """剪贴板 HTML 提取: Ctrl+A/C → 读 text/html → 解析 table（含 colspan/rowspan 合并单元格信息）"""
        try:
            await self._do_ctrl_a_c(page)

            # 读剪贴板 HTML 格式
            html_content = await page.evaluate("""async () => {
                try {
                    const items = await navigator.clipboard.read();
                    for (const item of items) {
                        if (item.types.includes('text/html')) {
                            const blob = await item.getType('text/html');
                            return await blob.text();
                        }
                    }
                    return null;
                } catch(e) {
                    return null;
                }
            }""")

            if not html_content or len(html_content) < 20:
                return {"success": False, "error": "剪贴板无 HTML 格式数据"}

            # 用 Python 解析 HTML table
            from html.parser import HTMLParser

            class TableParser(HTMLParser):
                def __init__(self):
                    super().__init__()
                    self.rows = []
                    self.current_row = []
                    self.current_cell = {"text": "", "colspan": 1, "rowspan": 1}
                    self.in_td = False
                    self.in_table = False

                def handle_starttag(self, tag, attrs):
                    attrs_dict = dict(attrs)
                    if tag == "table":
                        self.in_table = True
                    elif tag in ("td", "th") and self.in_table:
                        self.in_td = True
                        self.current_cell = {
                            "text": "",
                            "colspan": int(attrs_dict.get("colspan", 1)),
                            "rowspan": int(attrs_dict.get("rowspan", 1)),
                        }
                    elif tag == "tr" and self.in_table:
                        self.current_row = []
                    elif tag == "br" and self.in_td:
                        self.current_cell["text"] += "\n"

                def handle_endtag(self, tag):
                    if tag in ("td", "th") and self.in_td:
                        self.in_td = False
                        self.current_row.append(self.current_cell)
                    elif tag == "tr" and self.in_table:
                        if self.current_row:
                            self.rows.append(self.current_row)
                        self.current_row = []
                    elif tag == "table":
                        self.in_table = False

                def handle_data(self, data):
                    if self.in_td:
                        self.current_cell["text"] += data

            parser = TableParser()
            parser.feed(html_content)

            if not parser.rows:
                return {"success": False, "error": "HTML 中无 table 数据"}

            # 展开合并单元格为二维矩阵
            grid = {}  # (row, col) → value
            max_col = 0
            for r, row_cells in enumerate(parser.rows):
                col_offset = 0
                for cell in row_cells:
                    # 跳过已被 rowspan 占据的位置
                    while (r, col_offset) in grid:
                        col_offset += 1
                    text = cell["text"].strip()
                    cs, rs = cell["colspan"], cell["rowspan"]
                    for dr in range(rs):
                        for dc in range(cs):
                            grid[(r + dr, col_offset + dc)] = text
                    col_offset += cs
                    max_col = max(max_col, col_offset)

            if not grid:
                return {"success": False, "error": "解析后无单元格数据"}

            # 第 0 行作为表头
            headers = []
            for c in range(max_col):
                headers.append(grid.get((0, c), f"col_{c}"))

            # 数据行
            records = []
            max_row = max(r for r, _ in grid.keys()) if grid else 0
            for r in range(1, max_row + 1):
                row = {}
                for c in range(max_col):
                    key = headers[c] if c < len(headers) else f"col_{c}"
                    row[key] = grid.get((r, c), "")
                records.append(row)

            # 合并单元格信息（从 grid 构建中已保留）
            merged_info = []
            for r, row_cells in enumerate(parser.rows):
                col_offset = 0
                for cell in row_cells:
                    while (r, col_offset) in grid and cell["text"].strip() != grid.get((r, col_offset), ""):
                        col_offset += 1
                    if cell["colspan"] > 1 or cell["rowspan"] > 1:
                        merged_info.append({
                            "start_row": r, "start_col": col_offset,
                            "colspan": cell["colspan"], "rowspan": cell["rowspan"],
                            "value": cell["text"].strip(),
                        })
                    col_offset += cell["colspan"]

            return {
                "success": True,
                "doc_type": "e3",
                "path_type": "sheet",
                "url": page.url,
                "headers": headers,
                "total": len(records),
                "records": records,
                "merged_cells": merged_info,
                "merged_cell_count": len(merged_info),
            }
        except Exception as e:
            return {"success": False, "error": f"剪贴板 HTML 提取失败: {e}"}

    async def _try_xlsx_export_for_spreadsheet(self, page) -> dict:
        """xlsx 导出: 触发文档导出下载 → openpyxl 解析（完整保留合并单元格）
        
        ⚠️ 需要编辑权限。只读文档会失败并降级到下一策略。
        """
        import os
        try:
            # 用 JS 点击文件菜单按钮（绕过 Playwright viewport 检查）
            menu_clicked = await page.evaluate("""() => {
                const el = document.querySelector('#headerbar-filemenu')
                    || document.querySelector('[aria-label="按钮:文件操作"]')
                    || document.querySelector('#main-menu-file');
                if (el) { el.click(); return true; }
                return false;
            }""")
            if not menu_clicked:
                return {"success": False, "error": "未找到文件菜单按钮"}
            await page.wait_for_timeout(1500)

            # 点击"导出"菜单项
            export_clicked = await page.evaluate("""() => {
                const items = document.querySelectorAll('[role="menuitem"], [class*="menu-item"]');
                for (const item of items) {
                    const text = item.textContent?.trim();
                    if (item.offsetParent !== null && text === '导出') {
                        item.click();
                        return text;
                    }
                }
                return null;
            }""")
            if not export_clicked:
                return {"success": False, "error": "未找到导出选项（可能无编辑权限）"}
            await page.wait_for_timeout(1500)

            # 点击 xlsx 选项并等待下载
            xlsx_text = await page.evaluate("""() => {
                const items = document.querySelectorAll('[role="menuitem"], [class*="menu-item"], li');
                for (const item of items) {
                    const text = item.textContent?.trim();
                    if (item.offsetParent !== null && text && text.includes('.xlsx')) {
                        return text;
                    }
                }
                return null;
            }""")
            if not xlsx_text:
                return {"success": False, "error": "未找到 xlsx 导出选项"}

            # 触发下载
            async with page.expect_download(timeout=60000) as dl_info:
                await page.evaluate("""(text) => {
                    const items = document.querySelectorAll('[role="menuitem"], [class*="menu-item"], li');
                    for (const item of items) {
                        if (item.offsetParent !== null && item.textContent?.trim() === text) {
                            item.click();
                            return;
                        }
                    }
                }""", xlsx_text)

            download = await dl_info.value
            tmp_path = await download.path()
            if not tmp_path or not os.path.exists(tmp_path):
                return {"success": False, "error": "下载文件不存在"}

            # 用 openpyxl 解析
            import openpyxl
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            all_sheets = {}
            total_records = 0

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                merged_map = {}
                merged_ranges_list = []
                for mr in ws.merged_cells.ranges:
                    top_left_val = ws.cell(mr.min_row, mr.min_col).value
                    merged_ranges_list.append(str(mr))
                    for r in range(mr.min_row, mr.max_row + 1):
                        for c in range(mr.min_col, mr.max_col + 1):
                            merged_map[(r, c)] = top_left_val

                rows_data = []
                for row in ws.iter_rows(min_row=1, values_only=False):
                    row_dict = {}
                    for cell in row:
                        val = cell.value
                        if val is None and (cell.row, cell.column) in merged_map:
                            val = merged_map[(cell.row, cell.column)]
                        col_letter = cell.column_letter
                        row_dict[col_letter] = str(val) if val is not None else ""
                    rows_data.append(row_dict)

                all_sheets[sheet_name] = {
                    "rows": rows_data,
                    "row_count": len(rows_data),
                    "merged_ranges": merged_ranges_list,
                    "merged_cell_count": len(merged_map),
                }
                total_records += len(rows_data)

            try:
                os.unlink(tmp_path)
            except:
                pass

            return {
                "success": True,
                "doc_type": "e3",
                "path_type": "sheet",
                "url": page.url,
                "method": "xlsx-export",
                "sheets": all_sheets,
                "sheet_names": wb.sheetnames,
                "sheet_count": len(wb.sheetnames),
                "total": total_records,
            }
        except Exception as e:
            return {"success": False, "error": f"xlsx 导出失败: {e}"}

    async def _try_clipboard_for_spreadsheet(self, page) -> dict:
        """剪贴板提取（纯文本 TSV 兜底）: Ctrl+A/C → readText → 解析 TSV"""
        try:
            await self._do_ctrl_a_c(page)

            # 读取剪贴板纯文本
            clipboard_text = await page.evaluate(
                """async () => {
                    try {
                        return await navigator.clipboard.readText();
                    } catch(e) {
                        return null;
                    }
                }"""
            )

            if not clipboard_text or len(clipboard_text) < 10:
                return {"success": False, "error": "剪贴板内容为空"}

            # 解析 TSV
            lines = clipboard_text.strip().split("\n")
            if not lines:
                return {"success": False, "error": "剪贴板无数据行"}

            headers = lines[0].split("\t")
            rows = []
            for line in lines[1:]:
                cells = line.split("\t")
                row = {}
                for i, cell in enumerate(cells):
                    key = headers[i] if i < len(headers) else f"col_{i}"
                    row[key] = cell.strip()
                rows.append(row)

            return {
                "success": True,
                "doc_type": "e3",
                "path_type": "sheet",
                "url": page.url,
                "headers": headers,
                "total": len(rows),
                "records": rows,
                "warning": "剪贴板纯文本提取，合并单元格和多行文本可能丢失",
            }
        except Exception as e:
            return {"success": False, "error": f"剪贴板提取失败: {e}"}

    # ── 3c. 智能表格 dop-api ─────────────────────────────

    async def _read_smartsheet(
        self, user_id: str, url: str, sheet_id: str = None
    ) -> dict:
        """通过 dop-api 获取 s3_ 智能表格全量结构化数据。

        方案：拦截页面首次 get/sheet 请求获取 xsrf 等动态参数，
        然后用完整参数主动 fetch startrow=0 获取全量数据。
        smartsheet 字段是 base64+zlib 压缩格式，需在 Python 端解码。
        多子表：从 collab_client_vars 获取 workbook，遍历所有 smartsheet 子表。
        """
        from playwright.async_api import async_playwright

        state_file = str(self._state_file(user_id))
        info = parse_doc_url(url)

        if not info["doc_id"]:
            return {"success": False, "error": f"无法解析 URL: {url}"}
        if not sheet_id and info["tab"]:
            sheet_id = info["tab"]

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=True, args=["--no-sandbox", "--disable-setuid-sandbox"]
            )
            ctx = await browser.new_context(
                storage_state=state_file,
                viewport={"width": 1280, "height": 800},
            )
            page = await ctx.new_page()
            await _block_fonts(page)

            # 拦截页面首次 get/sheet 请求，提取 xsrf 等动态参数
            captured_params = None

            def _on_request(request):
                nonlocal captured_params
                if "dop-api/get/sheet" in request.url and not captured_params:
                    qs = parse_qs(urlparse(request.url).query)
                    captured_params = {k: v[0] for k, v in qs.items()}

            page.on("request", _on_request)

            try:
                await page.goto(url, wait_until="networkidle", timeout=60000)
            except:
                pass
            await page.wait_for_timeout(5000)

            if _is_login_page(page.url):
                await browser.close()
                return {
                    "success": False,
                    "error": "cookies 已过期，需重新登录",
                    "user_id": user_id,
                }

            if not captured_params:
                await browser.close()
                _auto_report("dop-api 未拦截到 get/sheet 请求",
                             "页面加载后未拦截到 dop-api/get/sheet 请求，无法获取 xsrf 参数。可能是页面结构变化或加载超时。",
                             {"url": url, "user_id": user_id, "step": "intercept_xsrf"})
                return {"success": False, "error": "未拦截到 dop-api/get/sheet 请求，无法获取 xsrf 参数"}

            xsrf = captured_params.get("xsrf", "")
            rev = captured_params.get("rev", "")
            doc_id = info["doc_id"]

            # 从 collab_client_vars 获取 workbook（子表列表）
            workbook = await page.evaluate("""() => {
                const ccv = window.clientVars?.collab_client_vars;
                if (!ccv) return null;
                const iat = ccv.initialAttributedText;
                if (!iat || !iat.text || !iat.text[0]) return null;
                const wb = iat.text[0].workbook;
                return wb ? JSON.parse(wb) : null;
            }""")

            if not workbook:
                await browser.close()
                return {"success": False, "error": "无法获取 workbook（子表列表）"}

            # 确定要读取的子表列表
            all_sheets = [s for s in workbook if s.get("type") == "smartsheet"]
            if sheet_id:
                # 只读指定子表
                target_sheets = [s for s in all_sheets if s["id"] == sheet_id]
                if not target_sheets:
                    target_sheets = [{"id": sheet_id, "name": sheet_id}]
            else:
                target_sheets = all_sheets

            # 对每个子表主动 fetch 全量数据
            all_results = []
            for sh in target_sheets:
                sid = sh["id"]
                sname = sh.get("name", sid)
                api_url = (
                    f"https://doc.weixin.qq.com/dop-api/get/sheet"
                    f"?padId={doc_id}&subId={sid}"
                    f"&startrow=0&endrow=99999&xsrf={xsrf}"
                    f"&outformat=1&normal=1&nowb=1&needSheetState=2"
                    f"&rev={rev}&optimizedVer=2"
                    f"&chunkCellSize=15000&enableChunkRank=1&enablePermOpt=0"
                )
                # JS 端返回原始 smartsheet 字符串（base64+zlib）
                fetch_result = await page.evaluate(
                    """async (url) => {
                        try {
                            const r = await fetch(url, { credentials: 'include' });
                            const j = await r.json();
                            const item = j?.data?.initialAttributedText?.text?.[0];
                            if (!item) return { error: j.errmsg || 'no text item', retcode: j.retcode };
                            return {
                                ok: true,
                                total_record_count: item.total_record_count,
                                max_row: item.max_row,
                                smartsheet: item.smartsheet,
                                workbook: item.workbook || '[]',
                            };
                        } catch (e) {
                            return { error: e.message };
                        }
                    }""",
                    api_url,
                )
                if not fetch_result or fetch_result.get("error"):
                    all_results.append({
                        "sheet_id": sid, "sheet_name": sname,
                        "success": False, "error": fetch_result.get("error") if fetch_result else "no response",
                    })
                    continue

                # Python 端解码 base64+zlib
                ss_raw = fetch_result["smartsheet"]
                try:
                    # 补齐 base64 padding
                    padding = 4 - len(ss_raw) % 4
                    if padding != 4:
                        ss_raw += "=" * padding
                    decoded = base64.urlsafe_b64decode(ss_raw)
                    decompressed = zlib.decompress(decoded)
                    parsed = json.loads(decompressed.decode("utf-8"))
                except Exception as e:
                    _auto_report(f"base64+zlib 解码失败 (sheet: {sname})",
                                 str(e),
                                 {"url": url, "sheet_id": sid, "sheet_name": sname, "step": "decode"})
                    all_results.append({
                        "sheet_id": sid, "sheet_name": sname,
                        "success": False, "error": f"base64+zlib 解码失败: {e}",
                    })
                    continue

                # parsed 可能是 [[items...]] 双层嵌套，取第一层
                if parsed and isinstance(parsed[0], list):
                    parsed = parsed[0]

                # 提取列定义和选项映射
                field_defs, select_options, user_map = _parse_column_defs(parsed)

                # 提取行数据（t=3028 的 item）
                records_raw = {}
                for item in parsed:
                    if item.get("t") == 3028:
                        c = item.get("c", {})
                        sub = c.get("2") or c.get("k2") or {}
                        records_raw = sub.get("1") or sub.get("k1") or {}
                        break

                if not records_raw:
                    all_results.append({
                        "sheet_id": sid, "sheet_name": sname,
                        "success": False, "error": "未找到行数据",
                    })
                    continue

                # 转换为结构化记录
                records = []
                for rid, rdata in records_raw.items():
                    cells = rdata.get("k1") or rdata.get("1") or {}
                    record = {"_record_id": rid}
                    for fid, fdef in field_defs.items():
                        cell = cells.get(fid, {})
                        so = select_options.get(fid) if fdef["type"] == _FIELD_TYPE_SELECT else None
                        record[fdef["name"]] = _cell_value(cell, so, user_map)
                    records.append(record)

                all_results.append({
                    "sheet_id": sid, "sheet_name": sname,
                    "success": True,
                    "records": records,
                    "field_defs": {fid: d["name"] for fid, d in field_defs.items()},
                    "total": len(records),
                    "total_record_count": fetch_result.get("total_record_count", len(records)),
                })

            await browser.close()

        # 单子表模式：直接返回该子表结果
        if len(all_results) == 1:
            r = all_results[0]
            if not r["success"]:
                return {"success": False, "error": r["error"], "sheet_id": r["sheet_id"]}
            return {
                "success": True,
                "doc_type": "s3",
                "records": r["records"],
                "field_defs": r["field_defs"],
                "sheet_id": r["sheet_id"],
                "sheet_name": r["sheet_name"],
                "workbook": workbook,
                "total": r["total"],
            }

        # 多子表模式：返回所有子表结果
        total_records = sum(r.get("total", 0) for r in all_results if r["success"])
        failed = [r for r in all_results if not r["success"]]
        return {
            "success": len(failed) == 0,
            "doc_type": "s3",
            "sheets": [
                {
                    "sheet_id": r["sheet_id"],
                    "sheet_name": r["sheet_name"],
                    "records": r.get("records", []),
                    "field_defs": r.get("field_defs", {}),
                    "total": r.get("total", 0),
                }
                for r in all_results if r["success"]
            ],
            "workbook": workbook,
            "total_sheets": len(all_results),
            "total_records": total_records,
            "failed_sheets": failed if failed else None,
        }

    # ── 3d. 思维导图 dop-api ──────────────────────────────

    async def _read_mind(self, user_id: str, url: str, info: dict) -> dict:
        """通过 dop-api/get/mind 获取 m4_ 思维导图完整节点树"""
        from playwright.async_api import async_playwright

        state_file = str(self._state_file(user_id))
        doc_id = info["doc_id"]

        if not doc_id:
            return {"success": False, "error": f"无法解析 doc_id: {url}"}

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=True, args=["--no-sandbox", "--disable-setuid-sandbox"]
            )
            ctx = await browser.new_context(
                storage_state=state_file,
                viewport={"width": 1280, "height": 800},
            )
            page = await ctx.new_page()
            await _block_fonts(page)

            try:
                await page.goto(url, wait_until="networkidle", timeout=30000)
            except:
                pass
            await page.wait_for_timeout(3000)

            if _is_login_page(page.url):
                await browser.close()
                return {
                    "success": False,
                    "error": "cookies 已过期，需重新登录",
                    "user_id": user_id,
                }

            # 调用 dop-api/get/mind
            api_url = (
                f"https://doc.weixin.qq.com/dop-api/get/mind"
                f"?padId={doc_id}&normal=1"
            )
            fetch_result = await page.evaluate(
                """async (url) => {
                    try {
                        const r = await fetch(url, { credentials: 'include' });
                        const j = await r.json();
                        // initialAttributedText.text 是 JSON 字符串（非数组）
                        const rawText = j?.data?.initialAttributedText?.text;
                        if (!rawText) return { error: 'no initialAttributedText.text' };
                        let parsed;
                        try {
                            parsed = JSON.parse(rawText);
                        } catch (e) {
                            return { error: 'text parse failed: ' + e.message, rawPreview: rawText.substring(0, 200) };
                        }
                        // 结构: { content: [{ rootTopic: {...} }] }
                        const contentArr = parsed?.content;
                        if (!contentArr || !contentArr.length) return { error: 'no content array', parsedKeys: Object.keys(parsed || {}) };
                        const rootTopic = contentArr[0]?.rootTopic;
                        if (!rootTopic) return { error: 'no rootTopic in content[0]', content0Keys: Object.keys(contentArr[0] || {}) };
                        return { ok: true, rootTopic: rootTopic };
                    } catch (e) {
                        return { error: e.message };
                    }
                }""",
                api_url,
            )
            await browser.close()

        if not fetch_result or fetch_result.get("error"):
            return {
                "success": False,
                "error": f"dop-api/get/mind 请求失败: {fetch_result}",
            }

        # 递归提取节点树
        root = fetch_result["rootTopic"]
        nodes = self._extract_mind_nodes(root)

        # 生成文本大纲
        text_lines = []
        for n in nodes:
            indent = "  " * n["depth"]
            text_lines.append(f"{indent}{n['title']}")

        return {
            "success": True,
            "doc_type": "m4",
            "path_type": info["path_type"],
            "title": root.get("title", ""),
            "total_nodes": len(nodes),
            "max_depth": max((n["depth"] for n in nodes), default=0),
            "nodes": nodes,
            "text": "\n".join(text_lines),
        }

    @staticmethod
    def _extract_mind_nodes(node, depth=0, result=None):
        """递归提取思维导图全部节点（适配 children.attached 结构）"""
        if result is None:
            result = []
        title = node.get("title", "")
        if title:
            result.append({
                "depth": depth,
                "title": title,
            })
        # 思维导图节点子节点在 children.attached 数组中
        children_container = node.get("children", {})
        attached = children_container.get("attached", []) if isinstance(children_container, dict) else []
        for child in attached:
            WeComDocReader._extract_mind_nodes(child, depth + 1, result)
        return result

    # ── 4. 用户管理 ──────────────────────────────────────

    def list_users(self) -> list:
        """列出所有已登录用户及其 cookies 状态"""
        users = []
        for f in sorted(self.state_dir.glob("*.json")):
            uid = f.stem
            remaining = None
            try:
                with open(f) as fp:
                    state = json.load(fp)
                for c in state.get("cookies", []):
                    if c["name"] == "wedoc_sid":
                        remaining = round(
                            (c.get("expires", 0) - time.time()) / 86400, 1
                        )
            except:
                pass
            users.append({
                "user_id": uid,
                "remaining_days": remaining,
                "expired": remaining is not None and remaining < 0,
            })
        return users

    def remove_user(self, user_id: str) -> dict:
        """删除指定用户的所有 cookies 和锁文件"""
        removed = []
        for f in [
            self._state_file(user_id),
            self._lock_file(user_id),
            self._qr_image(user_id),
        ]:
            if f.exists():
                f.unlink()
                removed.append(str(f))
        return {"removed": removed, "user_id": user_id}
